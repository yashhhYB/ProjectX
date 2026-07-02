"""
generate_xlsx.py — Generate professional XLSX submission with formatted output.

Creates a multi-sheet Excel workbook:
  Sheet 1: "Top 100 Ranking" — Main ranking with formatting
  Sheet 2: "Scoring Breakdown" — Detailed scores per component
  Sheet 3: "Methodology" — Brief explanation of approach

Usage:
    python generate_xlsx.py
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def generate_xlsx(results, output_path="submission.xlsx"):
    """Generate a professional XLSX file from ranking results."""

    wb = Workbook()

    # ═══════════════════════════════════════
    # Sheet 1: Top 100 Ranking
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Top 100 Ranking"

    # Styles
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    top10_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    top30_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title row
    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = "Redrob AI — Intelligent Candidate Discovery & Ranking (Top 100)"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[1].height = 35

    # Subtitle
    ws1.merge_cells("A2:F2")
    sub_cell = ws1["A2"]
    sub_cell.value = "Senior AI/ML Engineer — Ranking, Retrieval & Matching Systems"
    sub_cell.font = Font(name="Calibri", size=11, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 25

    # Headers (row 4)
    headers = ["Rank", "Candidate ID", "Score", "Title", "Experience (yr)", "Reasoning"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[4].height = 30

    # Data rows
    for i, entry in enumerate(results):
        row = i + 5
        candidate = entry.get("candidate", {})
        profile = candidate.get("profile", {})

        rank = entry["rank"]
        cid = entry["candidate_id"]
        score = entry["score"]
        title = profile.get("current_title", "Unknown")
        yoe = profile.get("years_of_experience", 0)
        reasoning = entry["reasoning"]

        ws1.cell(row=row, column=1, value=rank).font = Font(name="Calibri", size=10, bold=True)
        ws1.cell(row=row, column=2, value=cid).font = data_font
        score_cell = ws1.cell(row=row, column=3, value=score)
        score_cell.font = data_font
        score_cell.number_format = "0.000000"
        ws1.cell(row=row, column=4, value=title).font = data_font
        ws1.cell(row=row, column=5, value=round(yoe, 1)).font = data_font
        ws1.cell(row=row, column=6, value=reasoning).font = data_font

        # Row formatting
        for col in range(1, 7):
            cell = ws1.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 6))

            # Color-code by rank tier
            if rank <= 10:
                cell.fill = top10_fill
            elif rank <= 30:
                cell.fill = top30_fill
            elif i % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 15
    ws1.column_dimensions["F"].width = 80

    # Freeze panes
    ws1.freeze_panes = "A5"

    # ═══════════════════════════════════════
    # Sheet 2: Scoring Breakdown
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("Scoring Breakdown")

    # Title
    ws2.merge_cells("A1:L1")
    title_cell = ws2["A1"]
    title_cell.value = "Detailed Scoring Breakdown — 9-Component Composite Score"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws2.row_dimensions[1].height = 30

    # Headers
    score_headers = [
        "Rank", "Candidate ID", "Composite", "Title/Career (25%)",
        "Skill Match (18%)", "Semantic (12%)", "Experience (10%)",
        "Behavioral (10%)", "Career Depth (8%)", "Assessment (7%)",
        "Location (5%)", "Education (5%)"
    ]
    for col, header in enumerate(score_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[3].height = 40

    for i, entry in enumerate(results):
        row = i + 4
        scores = entry.get("scores_detail", {})

        ws2.cell(row=row, column=1, value=entry["rank"]).font = data_font
        ws2.cell(row=row, column=2, value=entry["candidate_id"]).font = data_font

        score_values = [
            scores.get("composite", 0),
            scores.get("title_career", 0),
            scores.get("skill_match", 0),
            scores.get("semantic", 0),
            scores.get("experience_fit", 0),
            scores.get("behavioral", 0),
            scores.get("career_depth", 0),
            scores.get("assessment", 0),
            scores.get("location", 0),
            scores.get("education", 0),
        ]

        for j, val in enumerate(score_values):
            cell = ws2.cell(row=row, column=j + 3, value=round(val, 4))
            cell.font = data_font
            cell.number_format = "0.0000"
            cell.border = thin_border

            # Color scale for scores
            if val >= 0.8:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif val >= 0.5:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif val < 0.3:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Style rank and ID
        for col in range(1, 3):
            ws2.cell(row=row, column=col).border = thin_border

        if i % 2 == 0:
            for col in range(1, 3):
                ws2.cell(row=row, column=col).fill = alt_fill

    # Column widths for sheet 2
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 18
    for col_idx in range(3, 13):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16
    ws2.freeze_panes = "A4"

    # ═══════════════════════════════════════
    # Sheet 3: Methodology
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("Methodology")

    methodology = [
        ("Redrob AI — Scoring Methodology", "", ""),
        ("", "", ""),
        ("Component", "Weight", "Description"),
        ("Title & Career", "25%", "Role relevance, seniority detection, ML career progression bonus"),
        ("Skill Match", "18%", "Log-weighted skill durability: proficiency x log(duration) x log(endorsements), core vs nice-to-have differentiation"),
        ("Semantic Similarity", "12%", "Cosine similarity using all-mpnet-base-v2 embeddings (768-dim) against enriched JD"),
        ("Experience Fit", "10%", "Gaussian distribution centered at 7yr (sigma=1.5) for 5-9yr sweet spot"),
        ("Behavioral Signals", "10%", "ALL 23 Redrob signals: response rate, recency, notice period, GitHub, verification, market signals"),
        ("Career Depth", "8%", "Product vs consulting ratio, production keyword density, career trajectory, certifications, industry alignment"),
        ("Skill Assessment", "7%", "Redrob platform assessment scores for verified competency"),
        ("Location", "5%", "Pune/Noida preference, India-based, willingness to relocate"),
        ("Education", "5%", "Institution tier, field relevance, degree level"),
        ("", "", ""),
        ("Key Design Decisions", "", ""),
        ("1. Honeypot Detection", "", "9 consistency checks: chronological impossibility, stuffed skills, domain mismatch, bot patterns, education timeline, expert inflation"),
        ("2. Job-Hopper Detection", "", "Avg tenure < 18 months triggers penalty (JD explicitly disqualifies title-chasers)"),
        ("3. Trust-Weighted Skills", "", "Skills scored as proficiency x log(duration) x log(endorsements) — penalizes keyword stuffing"),
        ("4. Behavioral as Multiplier", "", "Perfect-on-paper candidates with low engagement are down-weighted appropriately"),
        ("5. Unique Reasoning", "", "Each candidate gets substantively different reasoning referencing specific facts from their profile"),
    ]

    for i, (col1, col2, col3) in enumerate(methodology, 1):
        ws3.cell(row=i, column=1, value=col1)
        ws3.cell(row=i, column=2, value=col2)
        ws3.cell(row=i, column=3, value=col3)

        if i == 1:
            ws3.cell(row=i, column=1).font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
            ws3.merge_cells("A1:C1")
        elif i == 3 or i == 14:
            for col in range(1, 4):
                cell = ws3.cell(row=i, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
        elif 4 <= i <= 12 or 15 <= i <= 19:
            for col in range(1, 4):
                cell = ws3.cell(row=i, column=col)
                cell.font = data_font
                cell.border = thin_border
                if (i - 4) % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(wrap_text=True)

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 90

    # Save
    wb.save(output_path)
    print(f"  Professional XLSX written to: {output_path}")
    return output_path


if __name__ == "__main__":
    # If run standalone, load results from the pipeline
    from src.loader import load_candidates_jsonl
    from src.ranker import run_ranking_pipeline
    from src.config import CANDIDATES_FILE, EMBEDDINGS_DIR

    print("Running full pipeline on 100K candidates...")
    candidates = load_candidates_jsonl(CANDIDATES_FILE)

    results = run_ranking_pipeline(
        candidates=candidates,
        embeddings_dir=EMBEDDINGS_DIR,
        use_prefilter=True,
        output_path="submission.csv",
    )

    generate_xlsx(results, "submission.xlsx")
    print("\nDone! Files generated: submission.csv, submission.xlsx")
